import logging
import os
import asyncio

import openpyxl as pyxl
from openpyxl.styles import numbers
import pandas as pd
import numpy as np
from numba import jit

import work
from api import models, crud
from api.database import SessionLocal

logging.basicConfig(level=logging.WARNING, filename='logger.log', filemode='a',
                    format="%(asctime)s %(levelname)s %(message)s")

class ExcelHandler:
    """
    Класс для работы с Excel
    """
    def __init__(self, filename: str):
        self.filename = filename
        self.wb = pyxl.load_workbook(filename)
        self.sheets = list([])
        self.cur_sheet = 0

        if not self.wb.worksheets[self.cur_sheet]['A1'].value or not self.wb.worksheets[self.cur_sheet]['A1'].value != '':
            self.wb.active.title = str(self.cur_sheet+1)
            self.sheets.append(self.wb.active)
        while self.wb.worksheets[self.cur_sheet]['A1'].value and self.wb.worksheets[self.cur_sheet]['A1'].value != '':
            cell = self.wb.worksheets[self.cur_sheet]['A1']
            self.sheets.append(self.wb.worksheets[self.cur_sheet])
            self.cur_sheet += 1
            if self.cur_sheet == len(self.wb.worksheets):
                self.wb.create_sheet(str(self.cur_sheet + 1))
                self.sheets.append(self.wb.worksheets[self.cur_sheet])
        self.tools_uid = list([])
        self.db = SessionLocal()


    def saveWorkbook(self):
        self.wb.save(self.filename)

    async def writeStats(self, time_from: str, time_to: str, timeframe: str):
        """
        Метод для записи статистики в виде сводных таблиц в Excel
        :param time_from - время начала моделирования. Нужно для поиска нужного файла
        :param time_to - время конца моделирования. Нужно для поиска нужного файла
        :param timeframe - таймфрейм моделирования. Нужно для поиска нужного файла
        :return:
        """

        if not time_from or not time_to:
            logging.error("в методе ExcelHandler.writeStatus: не указаны границы периода моделирования")
            raise ValueError("в методе ExcelHandler.writeStatus: не указаны границы периода моделирования")
        if not timeframe:
            logging.error("в методе ExcelHandler.writeStatus: не указан таймфрейм моделирования")
            raise ValueError("в методе ExcelHandler.writeStatus: не указан таймфрейм моделирования")

        with open("instruments.txt") as tools_file:
            lines = tools_file.readlines()
            for line in lines:
                self.tools_uid.append(line.rstrip('\n'))

        self.sheets[self.cur_sheet]['A1'] = 'Риск для счета'
        self.sheets[self.cur_sheet]['B1'] = work.STOP_ACCOUNT
        self.sheets[self.cur_sheet]['A2'] = 'STOP_LOSS'
        self.sheets[self.cur_sheet]['B2'] = work.STOP_LOSS
        self.sheets[self.cur_sheet]['A3'] = 'TAKE_PROFIT'
        self.sheets[self.cur_sheet]['B3'] = 0
        self.sheets[self.cur_sheet]['A4'] = 'Интервал SMA'
        self.sheets[self.cur_sheet]['B4'] = work.SMA_INTERVAL
        self.sheets[self.cur_sheet]['A5'] = 'Интервал RSI'
        self.sheets[self.cur_sheet]['B5'] = work.RSI_INTERVAL
        self.sheets[self.cur_sheet]['A6'] = 'Таймфрейм'
        self.sheets[self.cur_sheet]['B6'] = timeframe
        self.sheets[self.cur_sheet]['A7'] = 'Время начала моделирования'
        self.sheets[self.cur_sheet]['B7'] = time_from
        self.sheets[self.cur_sheet]['A8'] = 'Время конца моделирования'
        self.sheets[self.cur_sheet]['B8'] = time_to
        self.sheets[self.cur_sheet]['A9'] = 'Начальное количество лотов в портфеле'
        self.sheets[self.cur_sheet]['B9'] = work.START_LOT_COUNT
        self.sheets[self.cur_sheet]['A10'] = 'Начальная сумма свободных денег в портфеле'
        self.sheets[self.cur_sheet]['B10'] = work.START_ACCOUNT_PORTFOLIO

        pathTrades = "../history_data/trades_stats"
        pathPrtf = "../history_data/portfolio_stats"
        if not os.path.exists(pathTrades):
            logging.error("Нет папки trades_stats")
            raise Exception("Нет папки trades_stats")
        if not os.path.exists(pathPrtf):
            logging.error("Нет папки portfolio_stats")
            raise Exception("Нет папки portfolio_stats")

        pathTrades = pathTrades + "/" + timeframe
        pathPrtf = pathPrtf + "/" + timeframe
        if not os.path.exists(pathTrades):
            logging.error(f"Нет папки trades_stats/{timeframe}")
            raise Exception(f"Нет папки trades_stats/{timeframe}")
        if not os.path.exists(pathPrtf):
            logging.error(f"Нет папки portfolio_stats/{timeframe}")
            raise Exception(f"Нет папки portfolio_stats/{timeframe}")


        self.sheets[self.cur_sheet]['D1'] = 'UID инструмента'
        self.sheets[self.cur_sheet]['D2'] = 'Имя инструмента'
        self.sheets[self.cur_sheet]['D3'] = 'Сектор'
        self.sheets[self.cur_sheet]['D4'] = 'Торговая площадка'
        self.sheets[self.cur_sheet]['D5'] = 'Тип инструмента'
        self.sheets[self.cur_sheet]['D6'] = 'START_FULL_PORTFOLIO, руб'
        self.sheets[self.cur_sheet]['D7'] = 'END_FULL_PORTFOLIO, руб'
        self.sheets[self.cur_sheet]['D8'] = 'FREE_MONEY_END, руб'
        self.sheets[self.cur_sheet]['D9'] = 'Количество лотов (конец)'
        self.sheets[self.cur_sheet]['D10'] = 'Прибыль/убыток, %'
        self.sheets[self.cur_sheet]['D11'] = 'Сумма сработтаных заявок по STOP-MARKET, руб'


        for i in range(len(self.tools_uid)):
            # Проходимся по каждому торговому инструменту с результатами моделирования и записываем в excel
            raw_tr_name = pathTrades + "/" + self.tools_uid[i] + "_" + time_from + "_" + time_to + ".csv"
            raw_prtf_name = pathPrtf + "/" + self.tools_uid[i] + "_" + time_from + "_" + time_to + ".csv"
            statsTradesFilename = raw_tr_name.replace(":", "_")
            statsPortfolioFilename = raw_prtf_name.replace(":", "_")
            if not os.path.isfile(statsTradesFilename) or not os.path.isfile(statsPortfolioFilename):
                logging.warning(f"Не существует файла с именем {statsTradesFilename}")
                logging.warning(f"Не существует файла с именем {statsPortfolioFilename}")
                continue

            await self.writeInstrument(i+5, statsPortfolioFilename, self.tools_uid[i])

    
    async def writeInstrument(self, column: int, prtf_file: str, uid: str):
        instrument = crud.get_instrument_uid(self.db, uid)
        name = instrument.name  # Название инструмента
        exchange = crud.get_exchange(self.db, instrument.exchange_id)
        exchange_name = exchange.name  # Название биржи
        sector = crud.get_sector(self.db, instrument.sector_id)
        sector_name = sector.name  # имя сектора
        instrument_type = crud.get_instrument_type(self.db, instrument.type_id)
        type_name = instrument_type.name
        dfPortfolio = pd.read_csv(prtf_file)

        if dfPortfolio.empty:
            return
        startFullSum = dfPortfolio['start_full_sum'].iloc[-1]
        endFullSum = dfPortfolio['cur_full_sum'].iloc[-1]
        profit = dfPortfolio['profit_in_percent'].iloc[-1]
        freeMoneyEnd = dfPortfolio['cur_free_sum'].iloc[-1]
        endCntLots = dfPortfolio['cur_cnt_lots'].iloc[-1]
        sumSM = dfPortfolio['SM_SELL_sum'].iloc[-1]

        self.sheets[self.cur_sheet].cell(row=1, column=column).value = uid            # UID инструмента
        self.sheets[self.cur_sheet].cell(row=2, column=column).value = name           # имя инструмента
        self.sheets[self.cur_sheet].cell(row=3, column=column).value = sector_name    # сектор
        self.sheets[self.cur_sheet].cell(row=4, column=column).value = exchange_name  # торговая площадка
        self.sheets[self.cur_sheet].cell(row=5, column=column).value = type_name      # тип инструмента
        self.sheets[self.cur_sheet].cell(row=6, column=column).value = startFullSum   # Начальная полная стоимость портфеля
        self.sheets[self.cur_sheet].cell(row=7, column=column).value = endFullSum     # Итоговая полная стоимость портфеля
        self.sheets[self.cur_sheet].cell(row=8, column=column).value = freeMoneyEnd   # Сумма свободных денег на момент конца торгов
        self.sheets[self.cur_sheet].cell(row=9, column=column).value = endCntLots     # Количество лотов на момент конца моделирования
        self.sheets[self.cur_sheet].cell(row=10, column=column).value = profit / 100  # Прибыль/убыток в процентах
        self.sheets[self.cur_sheet].cell(row=10, column=column).number_format = numbers.FORMAT_PERCENTAGE_00
        self.sheets[self.cur_sheet].cell(row=11, column=column).value = sumSM          # Сумма сработавших стоп-маркет ордеров

async def main():
    excelHandler = ExcelHandler("excel_stats.xlsx")
    task = asyncio.create_task(excelHandler.writeStats('2023-06-01_10:00:00', '2024-03-31_23:00:00', 'DAY'))
    res = await asyncio.gather(task)
    while not task.done():
        continue
    excelHandler.saveWorkbook()

if __name__ == '__main__':
    asyncio.run(main())